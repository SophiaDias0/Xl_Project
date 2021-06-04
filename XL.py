from openpyxl import Workbook

def CREATE_XL_FILE(filename, sheetname="XLSheet", titles=False, columns_number=10, lines_number=10, choice=0, choice2=0, numberx=0, labels=[''], first_row_val=[0], first_column_val=[0]):
    wb = Workbook()
    ws = wb.active
    ws.title = sheetname
    correctedValue = 0
    if titles:
        for c in range(1, columns_number + 1):
            cell = ws.cell(1, c)
            correctedCel = labels[c-1]
            cell.value = correctedCel
        if choice == 0: #according to each row
            for c in range(2, lines_number+1):
                cell = ws.cell(c, 1)
                cell.value = first_row_val[c-2]
            for c in range(2, lines_number+1):
                for i in range(2, columns_number+1):
                    cell = ws.cell(c, i)
                    if choice2 == 0:
                        if i == 2:
                            correctedValue = numberx + float(first_row_val[c-2])
                        else:
                            correctedValue += numberx
                        cell.value = correctedValue
                    elif choice2 == 1:
                        if i == 2:
                            correctedValue = float(first_row_val[c-2]) - numberx
                        else:
                            correctedValue -= numberx
                        cell.value = correctedValue
                    elif choice2 == 2:
                        if i == 2:
                            correctedValue = float(first_row_val[c-2]) * numberx
                        else:
                            correctedValue *= numberx
                        cell.value = correctedValue
                    elif choice2 == 3:
                        if i == 2:
                            correctedValue = float(first_row_val[c-2]) / numberx
                        else:
                            correctedValue /= numberx
                        cell.value = correctedValue
                    elif choice2 == 4:
                        if i == 2:
                            correctedValue = float(first_row_val[c-2]) ** numberx
                        else:
                            correctedValue **= numberx
                        cell.value = correctedValue
                    else:
                        if i == 2:
                            correctedValue = float(first_row_val[c-2]) ** 2
                        else:
                            correctedValue *= float(first_row_val[c-2])
                        cell.value = correctedValue
        else: #according to each column
            for c in range(1, columns_number+1):
                cell = ws.cell(2, c)
                cell.value = first_column_val[c-1]
            for c in range(1, columns_number+1):
                for i in range(3, lines_number+1):
                    cell = ws.cell(i, c)
                    if choice2 == 0:
                        if i == 3:
                            correctedValue = numberx + float(first_column_val[c-1])
                        else:
                            correctedValue += numberx
                        cell.value = correctedValue
                    elif choice2 == 1:
                        if i == 3:
                            correctedValue = float(first_column_val[c-1]) - numberx
                        else:
                            correctedValue -= numberx
                        cell.value = correctedValue
                    elif choice2 == 2:
                        if i == 3:
                            correctedValue = float(first_column_val[c-1]) * numberx
                        else:
                            correctedValue *= numberx
                        cell.value = correctedValue
                    elif choice2 == 3:
                        if i == 3:
                            correctedValue = float(first_column_val[c-1]) / numberx
                        else:
                            correctedValue /= numberx
                        cell.value = correctedValue
                    elif choice2 == 4:
                        if i == 3:
                            correctedValue = float(first_column_val[c-1]) ** numberx
                        else:
                            correctedValue **= numberx
                        cell.value = correctedValue
                    else:
                        if i == 3:
                            correctedValue = float(first_column_val[c-1]) ** 2
                        else:
                            correctedValue *= float(first_column_val[c-1])
                        cell.value = correctedValue
    else:
        if choice == 0: #according to each row
            for c in range(1, lines_number+1):
                cell = ws.cell(c, 1)
                cell.value = first_row_val[c-1]
            for c in range(1, lines_number+1):
                for i in range(2, columns_number+1):
                    cell = ws.cell(c, i)
                    if choice2 == 0:
                        if i == 2:
                            correctedValue = numberx + float(first_row_val[c-1])
                        else:
                            correctedValue += numberx
                        cell.value = correctedValue
                    elif choice2 == 1:
                        if i == 2:
                            correctedValue = float(first_row_val[c-1]) - numberx
                        else:
                            correctedValue -= numberx
                        cell.value = correctedValue
                    elif choice2 == 2:
                        if i == 2:
                            correctedValue = float(first_row_val[c-1]) * numberx
                        else:
                            correctedValue *= numberx
                        cell.value = correctedValue
                    elif choice2 == 3:
                        if i == 2:
                            correctedValue = float(first_row_val[c-1]) / numberx
                        else:
                            correctedValue /= numberx
                        cell.value = correctedValue
                    elif choice2 == 4:
                        if i == 2:
                            correctedValue = float(first_row_val[c-1]) ** numberx
                        else:
                            correctedValue **= numberx
                        cell.value = correctedValue
                    else:
                        if i == 2:
                            correctedValue = float(first_row_val[c-1]) ** 2
                        else:
                            correctedValue *= float(first_row_val[c-1])
                        cell.value = correctedValue
        else: #according to each column
            for c in range(1, columns_number+1):
                cell = ws.cell(1, c)
                cell.value = first_column_val[c-1]
            for c in range(1, columns_number+1):
                for i in range(2, lines_number+1):
                    cell = ws.cell(i, c)
                    if choice2 == 0:
                        if i == 2:
                            correctedValue = numberx + float(first_column_val[c-1])
                        else:
                            correctedValue += numberx
                        cell.value = correctedValue
                    elif choice2 == 1:
                        if i == 2:
                            correctedValue = float(first_column_val[c-1]) - numberx
                        else:
                            correctedValue -= numberx
                        cell.value = correctedValue
                    elif choice2 == 2:
                        if i == 2:
                            correctedValue = float(first_column_val[c-1]) * numberx
                        else:
                            correctedValue *= numberx
                        cell.value = correctedValue
                    elif choice2 == 3:
                        if i == 2:
                            correctedValue = float(first_column_val[c-1]) / numberx
                        else:
                            correctedValue /= numberx
                        cell.value = correctedValue
                    elif choice2 == 4:
                        if i == 2:
                            correctedValue = float(first_column_val[c-1]) ** numberx
                        else:
                            correctedValue **= numberx
                        cell.value = correctedValue
                    else:
                        if i == 2:
                            correctedValue = float(first_column_val[c-1]) ** 2
                        else:
                            correctedValue *= float(first_column_val[c-1])
                        cell.value = correctedValue
    wb.save(filename)
